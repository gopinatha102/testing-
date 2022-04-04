import openpyxl


class Home_Page_Data:
    Home_Page_Data = [{"user_name": "  ", "password": "secret_sauce"}, {"user_name": "standard_user", "password": " "}, {"user_name": "standard_user", "password": "secret_sa"},{"user_name":"standard_user","password":"secret_sauce"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\DELL\\OneDrive\\Desktop\\demo.xlsx")
        Dict = {}
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]